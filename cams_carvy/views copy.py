# views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import CamsPortfolio,CAMSTransaction, CAMSSummary
import json
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404
import csv
from django.views import View
from .services import process_cams_files,import_cams_portfolio_with_pandas,process_cams_emails
from django.views.decorators.http import require_GET,require_POST
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View
from django.db.models import Q
from .models import CamsPortfolio  # Adjust import as needed

from django.views import View
from django.http import HttpResponse
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from .models import CamsPortfolio  # Replace 'your_app' with your actual app name

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q, Count
from decimal import Decimal, InvalidOperation

@login_required
def cams_carvy_dashboard(request):
    """Main dashboard view"""
    # Get all portfolios for summary stats
    all_portfolios = CamsPortfolio.objects.all()
    summary_stats = CamsPortfolio.get_summary_stats(all_portfolios)
    
    context = {
        'summary_stats': summary_stats,
    }
    return render(request, 'test.html', context)

# @login_required
# def get_portfolio_data(request):
#     """Ajax endpoint for portfolio data with filtering, pagination, and search"""
#     # Get query parameters
#     draw = int(request.GET.get('draw', 1))
#     start = int(request.GET.get('start', 0))
#     length = int(request.GET.get('length', 10))
#     search_value = request.GET.get('search[value]', '')
    
#     # Column search parameters
#     entity_filter = request.GET.get('entity_filter', '')
#     isin_filter = request.GET.get('isin_filter', '')
#     scheme_filter = request.GET.get('scheme_filter', '')
#     month_filter = request.GET.get('month_filter', '')
    
#     # Individual column searches
#     folio_search = request.GET.get('col_0_search', '')
#     entity_search = request.GET.get('col_1_search', '')
#     isin_search = request.GET.get('col_2_search', '')
#     scheme_search = request.GET.get('col_3_search', '')
#     cost_search = request.GET.get('col_4_search', '')
#     unit_search = request.GET.get('col_5_search', '')
#     nav_search = request.GET.get('col_6_search', '')
#     market_search = request.GET.get('col_7_search', '')
#     updated_search = request.GET.get('col_8_search', '')
#     user_search = request.GET.get('col_9_search', '')
    
#     # Start with all portfolios
#     portfolios = CamsPortfolio.objects.all()
    
#     # Apply filters
#     if entity_filter:
#         portfolios = portfolios.filter(entity_name__icontains=entity_filter)
#     if isin_filter:
#         portfolios = portfolios.filter(isin__icontains=isin_filter)
#     if scheme_filter:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
#     if month_filter:
#         portfolios = portfolios.filter(month=month_filter)
    
#     # Apply global search
#     if search_value:
#         portfolios = portfolios.filter(
#             Q(folio_no__icontains=search_value) |
#             Q(entity_name__icontains=search_value) |
#             Q(isin__icontains=search_value) |
#             Q(scheme_name__icontains=search_value)
#         )
    
#     # Apply column-specific searches
#     if folio_search:
#         portfolios = portfolios.filter(folio_no__icontains=folio_search)
#     if entity_search:
#         portfolios = portfolios.filter(entity_name__icontains=entity_search)
#     if isin_search:
#         portfolios = portfolios.filter(isin__icontains=isin_search)
#     if scheme_search:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
#     if cost_search:
#         portfolios = portfolios.filter(cost_value__icontains=cost_search)
#     if unit_search:
#         portfolios = portfolios.filter(unit_balance__icontains=unit_search)
#     if nav_search:
#         portfolios = portfolios.filter(nav_date__icontains=nav_search)
#     if market_search:
#         portfolios = portfolios.filter(market_value__icontains=market_search)
#     if updated_search:
#         portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
#     if user_search:
#         portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
#     # Get total count before pagination
#     total_records = portfolios.count()
    
#     # Apply ordering (default by updated_at desc)
#     order_column = request.GET.get('order[0][column]', '8')  # Default to updated_at
#     order_dir = request.GET.get('order[0][dir]', 'desc')
    
#     order_columns = {
#         '0': 'folio_no',
#         '1': 'entity_name', 
#         '2': 'isin',
#         '3': 'scheme_name',
#         '4': 'cost_value',
#         '5': 'unit_balance',
#         '6': 'nav_date',
#         '7': 'market_value',
#         '8': 'updated_at',
#         '9': 'updated_by__username'
#     }
    
#     order_field = order_columns.get(order_column, 'updated_at')
#     if order_dir == 'desc':
#         order_field = f'-{order_field}'
    
#     portfolios = portfolios.select_related('updated_by').order_by(order_field)
    
#     # Apply pagination
#     portfolios = portfolios[start:start + length]
    
#     # Prepare data for DataTables
#     data = []
#     for portfolio in portfolios:
#         profit_loss = portfolio.profit_loss
#         is_profit = profit_loss >= 0
#         profit_class = 'text-success' if is_profit else 'text-danger'
#         profit_icon = 'bi-arrow-up-right' if is_profit else 'bi-arrow-down-right'
        
#         data.append([
#             f'<a href="#" class="fw-bold text-primary">{portfolio.folio_no}</a>',
#             portfolio.entity_name,
#             f'<span class="editable-isin" data-id="{portfolio.id}" data-value="{portfolio.isin}">{portfolio.isin}</span>',
#             portfolio.scheme_name,
#             f'₹{portfolio.cost_value:,.2f}',
#             f'{portfolio.unit_balance:,.3f}',
#             portfolio.nav_date.strftime('%Y-%m-%d'),
#             # f'<span class="{profit_class} fw-bold">₹{portfolio.market_value:,.2f} <i class="bi {profit_icon}"></i></span>',
#             # portfolio.updated_at.strftime('%Y-%m-%d'),
#             # portfolio.updated_by.username if portfolio.updated_by else 'System'
#             f'<span class="{profit_class} fw-bold">₹{portfolio.market_value:,.2f} <i class="bi {profit_icon}"></i></span>',
#             portfolio.updated_at.strftime('%Y-%m-%d') if portfolio.updated_at else 'NA',
#             portfolio.updated_by.username if portfolio.updated_by else 'NA',

#         ])
    
#     return JsonResponse({
#         'draw': draw,
#         'recordsTotal': CamsPortfolio.objects.count(),
#         'recordsFiltered': total_records,
#         'data': data
#     })



# @csrf_exempt
# @login_required
# def update_isin(request):
#     """Ajax endpoint to update ISIN"""
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             portfolio_id = data.get('id')
#             new_isin = data.get('isin', '').strip()
            
#             if not new_isin:
#                 return JsonResponse({'success': False, 'error': 'ISIN cannot be empty'})
            
#             if len(new_isin) != 12:
#                 return JsonResponse({'success': False, 'error': 'ISIN must be 12 characters long'})
            
#             portfolio = get_object_or_404(CamsPortfolio, id=portfolio_id)
#             portfolio.isin = new_isin
#             portfolio.updated_by = request.user
#             portfolio.save()
            
#             return JsonResponse({'success': True, 'isin': new_isin})
            
#         except Exception as e:
#             return JsonResponse({'success': False, 'error': str(e)})
    
#     return JsonResponse({'success': False, 'error': 'Invalid request method'})

def format_indian_number(number, decimals=2):
    """
    Convert a number to Indian style format with commas.
    Example: 490028384.45 -> 49,00,28,384.45
    """
    if number is None:
        return f"0.{ '0'*decimals }"

    try:
        number = float(number)
    except (ValueError, TypeError):
        return str(number)

    # Split integer and decimal parts
    int_part, dec_part = divmod(number, 1)
    dec_part = round(dec_part * (10 ** decimals))  # handle decimals

    int_part_str = str(int(int_part))

    if len(int_part_str) <= 3:
        formatted_int = int_part_str
    else:
        last3 = int_part_str[-3:]
        rest = int_part_str[:-3]
        rest_groups = []
        while len(rest) > 2:
            rest_groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            rest_groups.insert(0, rest)
        formatted_int = ','.join(rest_groups) + ',' + last3

    return f"{formatted_int}.{dec_part:0{decimals}d}"





@login_required
def get_portfolio_data(request):
    """Ajax endpoint for portfolio data with filtering, pagination, and search"""
    # Get query parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    # Column search parameters
    entity_filter = request.GET.get('entity_filter', '')
    isin_filter = request.GET.get('isin_filter', '')
    scheme_filter = request.GET.get('scheme_filter', '')
    month_filter = request.GET.get('month_filter', '')
    
    # Individual column searches - UPDATED MAPPING
    sno_search = request.GET.get('col_0_search', '')        # S.No (not used for filtering)
    entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped position)
    folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped position)
    isin_search = request.GET.get('col_3_search', '')       # ISIN
    scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
    cost_search = request.GET.get('col_5_search', '')       # Cost Value
    unit_search = request.GET.get('col_6_search', '')       # Unit Balance
    nav_search = request.GET.get('col_7_search', '')        # NAV Date
    market_search = request.GET.get('col_8_search', '')     # Market Value
    updated_search = request.GET.get('col_9_search', '')    # Updated On
    user_search = request.GET.get('col_10_search', '')      # Updated By
    
    # Start with all portfolios
    portfolios = CamsPortfolio.objects.all()
    
    # Apply filters
    if entity_filter:
        portfolios = portfolios.filter(entity_name__icontains=entity_filter)
    if isin_filter:
        portfolios = portfolios.filter(isin__icontains=isin_filter)
    if scheme_filter:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
    if month_filter:
        portfolios = portfolios.filter(month=month_filter)
    
    # Apply global search
    if search_value:
        portfolios = portfolios.filter(
            Q(folio_no__icontains=search_value) |
            Q(entity_name__icontains=search_value) |
            Q(isin__icontains=search_value) |
            Q(scheme_name__icontains=search_value)
        )
    
    # Apply column-specific searches
    if folio_search:
        portfolios = portfolios.filter(folio_no__icontains=folio_search)
    if entity_search:
        portfolios = portfolios.filter(entity_name__icontains=entity_search)
    if isin_search:
        portfolios = portfolios.filter(isin__icontains=isin_search)
    if scheme_search:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
    if cost_search:
        portfolios = portfolios.filter(cost_value__icontains=cost_search)
    if unit_search:
        portfolios = portfolios.filter(unit_balance__icontains=unit_search)
    if nav_search:
        portfolios = portfolios.filter(nav_date__icontains=nav_search)
    if market_search:
        portfolios = portfolios.filter(market_value__icontains=market_search)
    if updated_search:
        portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
    if user_search:
        portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
    # Get total count before pagination
    total_records = portfolios.count()
    
    # Apply ordering - UPDATED COLUMN MAPPING
    order_column = request.GET.get('order[0][column]', '9')  # Default to updated_at (column 9)
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    order_columns = {
        '0': 'id',                      # S.No - use ID for ordering
        '1': 'entity_name',             # Entity Name (swapped)
        '2': 'folio_no',                # Folio No (swapped)
        '3': 'isin',                    # ISIN
        '4': 'scheme_name',             # Scheme Name
        '5': 'cost_value',              # Cost Value
        '6': 'unit_balance',            # Unit Balance
        '7': 'nav_date',                # NAV Date
        '8': 'market_value',            # Market Value
        '9': 'updated_at',              # Updated On
        '10': 'updated_by__username'    # Updated By
    }
    
    order_field = order_columns.get(order_column, 'updated_at')
    if order_dir == 'desc':
        order_field = f'-{order_field}'
    
    portfolios = portfolios.select_related('updated_by').order_by(order_field)
    
    # Apply pagination
    portfolios = portfolios[start:start + length]
    
    # Prepare data for DataTables with S.No - UPDATED DATA ORDER
    data = []
    for index, portfolio in enumerate(portfolios):
        # print(type(portfolio.market_value),"'////")
        # if ',' in portfolio.market_value :
        #         market_value =  portfolio.market_value.replace(',','')

        profit_loss = portfolio.profit_loss
        is_profit = profit_loss >= 0
        profit_class = 'text-success' if is_profit else 'text-danger'
        profit_icon = 'bi-arrow-up-right' if is_profit else 'bi-arrow-down-right'
        
        # Calculate S.No based on pagination
        sno = start + index + 1
        
        data.append([
            sno,  # Column 0: S.No
            portfolio.entity_name,  # Column 1: Entity Name (swapped)
            f'<span href="#" class=" ">{portfolio.folio_no}</span>',  # Column 2: Folio No (swapped)
            f'<span class="editable-isin" data-id="{portfolio.id}" data-value="{portfolio.isin}">{portfolio.isin}</span>',  # Column 3: ISIN
            portfolio.scheme_name,  # Column 4: Scheme Name
            
            # f'₹{portfolio.cost_value:,.2f}',  # Column 5: Cost Value
            f'₹{format_indian_number(str(portfolio.cost_value).replace(",",""))}',  # Column 5: Cost Value
            # f'₹{portfolio.unit_balance:,.2f}',  # Column 6: Unit Balance
            f'₹{format_indian_number(str(portfolio.unit_balance).replace(",",""))}',  # Column 6: Unit Balance

            portfolio.nav_date.strftime('%d-%m-%Y'),  # Column 7: NAV Date
            
            # f'<span class="{profit_class} ">₹{portfolio.market_value:,.2f}',  # Column 8: Market Value
            f'<span  ">₹{format_indian_number(str(portfolio.market_value))}',  # Column 8: Market Value
            
            
            portfolio.updated_at.strftime('%d-%m-%Y') if portfolio.updated_at else 'NA',  # Column 9: Updated On
            portfolio.updated_by.username if portfolio.updated_by else 'NA',  # Column 10: Updated By
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': CamsPortfolio.objects.count(),
        'recordsFiltered': total_records,
        'data': data
    })


# @login_required
# def get_summary_stats(request):
#     """Ajax endpoint for summary statistics with filtering"""
    
#     # Get all the same filter parameters as get_portfolio_data - UPDATED MAPPING
#     entity_filter = request.GET.get('entity_filter', '')
#     isin_filter = request.GET.get('isin_filter', '')
#     scheme_filter = request.GET.get('scheme_filter', '')
#     month_filter = request.GET.get('month_filter', '')
    
#     # Global search
#     global_search = request.GET.get('global_search', '')
    
#     # Individual column searches - UPDATED MAPPING
#     sno_search = request.GET.get('col_0_search', '')        # S.No (not used)
#     entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped)
#     folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped)
#     isin_search = request.GET.get('col_3_search', '')       # ISIN
#     scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
#     cost_search = request.GET.get('col_5_search', '')       # Cost Value
#     unit_search = request.GET.get('col_6_search', '')       # Unit Balance
#     nav_search = request.GET.get('col_7_search', '')        # NAV Date
#     market_search = request.GET.get('col_8_search', '')     # Market Value
#     updated_search = request.GET.get('col_9_search', '')    # Updated On
#     user_search = request.GET.get('col_10_search', '')      # Updated By
    
#     # Start with all portfolios - SAME FILTERING LOGIC AS get_portfolio_data
#     portfolios = CamsPortfolio.objects.all()
    
#     # Apply main filters
#     if entity_filter:
#         portfolios = portfolios.filter(entity_name__icontains=entity_filter)
#     if isin_filter:
#         portfolios = portfolios.filter(isin__icontains=isin_filter)
#     if scheme_filter:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
#     if month_filter:
#         portfolios = portfolios.filter(month=month_filter)
    
#     # Apply global search
#     if global_search:
#         portfolios = portfolios.filter(
#             Q(folio_no__icontains=global_search) |
#             Q(entity_name__icontains=global_search) |
#             Q(isin__icontains=global_search) |
#             Q(scheme_name__icontains=global_search)
#         )
    
#     # Apply column-specific searches
#     if folio_search:
#         portfolios = portfolios.filter(folio_no__icontains=folio_search)
#     if entity_search:
#         portfolios = portfolios.filter(entity_name__icontains=entity_search)
#     if isin_search:
#         portfolios = portfolios.filter(isin__icontains=isin_search)
#     if scheme_search:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
#     if cost_search:
#         portfolios = portfolios.filter(cost_value__icontains=cost_search)
#     if unit_search:
#         portfolios = portfolios.filter(unit_balance__icontains=unit_search)
#     if nav_search:
#         portfolios = portfolios.filter(nav_date__icontains=nav_search)
#     if market_search:
#         portfolios = portfolios.filter(market_value__icontains=market_search)
#     if updated_search:
#         portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
#     if user_search:
#         portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
#     # Calculate statistics on the filtered queryset
#     stats = portfolios.aggregate(
#         total_market_value=Sum('market_value'),
#         total_cost_value=Sum('cost_value'),
#         total_folios=Count('id'),
#         total_entities=Count('entity_name', distinct=True)
#     )
    
#     return JsonResponse({
#         'total_entities': stats['total_entities'] or 0,
#         'total_folios': stats['total_folios'] or 0,
#         'total_market_value': float(stats['total_market_value'] or 0),
#         'total_cost_value': float(stats['total_cost_value'] or 0)
#     })





@login_required
def get_summary_stats(request):
    """Ajax endpoint for summary statistics with filtering"""
    
    # Get all the same filter parameters as get_portfolio_data - UPDATED MAPPING
    entity_filter = request.GET.get('entity_filter', '')
    isin_filter = request.GET.get('isin_filter', '')
    scheme_filter = request.GET.get('scheme_filter', '')
    month_filter = request.GET.get('month_filter', '')
    
    # Global search
    global_search = request.GET.get('global_search', '')
    
    # Individual column searches - UPDATED MAPPING
    sno_search = request.GET.get('col_0_search', '')        # S.No (not used)
    entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped)
    folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped)
    isin_search = request.GET.get('col_3_search', '')       # ISIN
    scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
    cost_search = request.GET.get('col_5_search', '')       # Cost Value
    unit_search = request.GET.get('col_6_search', '')       # Unit Balance
    nav_search = request.GET.get('col_7_search', '')        # NAV Date
    market_search = request.GET.get('col_8_search', '')     # Market Value
    updated_search = request.GET.get('col_9_search', '')    # Updated On
    user_search = request.GET.get('col_10_search', '')      # Updated By
    
    # Start with all portfolios - SAME FILTERING LOGIC AS get_portfolio_data
    portfolios = CamsPortfolio.objects.all()
    
    # Apply main filters
    if entity_filter:
        portfolios = portfolios.filter(entity_name__icontains=entity_filter)
    if isin_filter:
        portfolios = portfolios.filter(isin__icontains=isin_filter)
    if scheme_filter:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
    if month_filter:
        portfolios = portfolios.filter(month=month_filter)
    
    # Apply global search
    if global_search:
        portfolios = portfolios.filter(
            Q(folio_no__icontains=global_search) |
            Q(entity_name__icontains=global_search) |
            Q(isin__icontains=global_search) |
            Q(scheme_name__icontains=global_search)
        )
    
    # Apply column-specific searches
    if folio_search:
        portfolios = portfolios.filter(folio_no__icontains=folio_search)
    if entity_search:
        portfolios = portfolios.filter(entity_name__icontains=entity_search)
    if isin_search:
        portfolios = portfolios.filter(isin__icontains=isin_search)
    if scheme_search:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
    if cost_search:
        portfolios = portfolios.filter(cost_value__icontains=cost_search)
    if unit_search:
        portfolios = portfolios.filter(unit_balance__icontains=unit_search)
    if nav_search:
        portfolios = portfolios.filter(nav_date__icontains=nav_search)
    if market_search:
        portfolios = portfolios.filter(market_value__icontains=market_search)
    if updated_search:
        portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
    if user_search:
        portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
    # Calculate statistics manually since CharField can't use Sum()
    total_market_value = Decimal('0')
    total_cost_value = Decimal('0')
    
    for portfolio in portfolios:
        try:
            if portfolio.market_value:
                market_val = Decimal(str(portfolio.market_value).replace(',', ''))
                total_market_value += market_val
        except (ValueError, InvalidOperation, TypeError):
            pass
            
        try:
            if portfolio.cost_value:
                cost_val = Decimal(str(portfolio.cost_value).replace(',', ''))
                total_cost_value += cost_val
        except (ValueError, InvalidOperation, TypeError):
            pass
    
    # Get count statistics using Django ORM
    count_stats = portfolios.aggregate(
        total_folios=Count('id'),
        total_entities=Count('entity_name', distinct=True)
    )
    
    return JsonResponse({
        'total_entities': count_stats['total_entities'] or 0,
        'total_folios': count_stats['total_folios'] or 0,
        'total_market_value': float(total_market_value),
        'total_cost_value': float(total_cost_value)
    })


@csrf_exempt
@login_required
def update_isin(request):
    """Ajax endpoint to update ISIN with updated validation (8-20 characters)"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            portfolio_id = data.get('id')
            new_isin = data.get('isin', '').strip()
            
            if not new_isin:
                return JsonResponse({'success': False, 'error': 'ISIN cannot be empty'})
            
            # UPDATED: Changed validation from 12 chars to 8-20 chars
            if len(new_isin) < 8 or len(new_isin) > 20:
                return JsonResponse({'success': False, 'error': 'ISIN must be 8 to 20 characters long'})
            
            portfolio = get_object_or_404(CamsPortfolio, id=portfolio_id)
            portfolio.isin = new_isin
            portfolio.updated_by = request.user
            portfolio.save()
            
            return JsonResponse({'success': True, 'isin': new_isin})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})





class DownloadSummaryView(View):
    def get(self, request):
        # Filters
        entity_filter = request.GET.get('entity_filter', '')
        isin_filter = request.GET.get('isin_filter', '')
        scheme_filter = request.GET.get('scheme_filter', '')
        month_filter = request.GET.get('month_filter', '')
        global_search = request.GET.get('global_search', '')

        # Column-specific searches
        sno_search = request.GET.get('col_0_search', '')
        entity_search = request.GET.get('col_1_search', '')
        folio_search = request.GET.get('col_2_search', '')
        isin_search = request.GET.get('col_3_search', '')
        scheme_search = request.GET.get('col_4_search', '')
        cost_search = request.GET.get('col_5_search', '')
        unit_search = request.GET.get('col_6_search', '')
        nav_search = request.GET.get('col_7_search', '')
        market_search = request.GET.get('col_8_search', '')
        updated_search = request.GET.get('col_9_search', '')
        user_search = request.GET.get('col_10_search', '')

        # Queryset
        portfolios = CamsPortfolio.objects.all()

        if entity_filter:
            portfolios = portfolios.filter(entity_name__icontains=entity_filter)
        if isin_filter:
            portfolios = portfolios.filter(isin__icontains=isin_filter)
        if scheme_filter:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
        if month_filter:
            portfolios = portfolios.filter(month=month_filter)

        if global_search:
            portfolios = portfolios.filter(
                Q(folio_no__icontains=global_search) |
                Q(entity_name__icontains=global_search) |
                Q(isin__icontains=global_search) |
                Q(scheme_name__icontains=global_search)
            )

        # Column-specific filtering
        if folio_search:
            portfolios = portfolios.filter(folio_no__icontains=folio_search)
        if entity_search:
            portfolios = portfolios.filter(entity_name__icontains=entity_search)
        if isin_search:
            portfolios = portfolios.filter(isin__icontains=isin_search)
        if scheme_search:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
        if cost_search:
            portfolios = portfolios.filter(cost_value__icontains=cost_search)
        if unit_search:
            portfolios = portfolios.filter(unit_balance__icontains=unit_search)
        if nav_search:
            portfolios = portfolios.filter(nav_date__icontains=nav_search)
        if market_search:
            portfolios = portfolios.filter(market_value__icontains=market_search)
        if updated_search:
            portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
        if user_search:
            portfolios = portfolios.filter(updated_by__username__icontains=user_search)

        portfolios = portfolios.select_related('updated_by').order_by('-updated_at')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CAMS Summary"

        # Header
        headers = [
            'S.No',
            'Entity Name',
            'Folio No',
            'ISIN',
            'Scheme Name',
            'Cost Value',
            'Unit Balance',
            'NAV Date',
            'Market Value',
            'Profit/Loss',
            'Updated On',
            'Updated By'
        ]
        ws.append(headers)

        # Helper function to safely convert string to float
        def safe_float_conversion(value, decimal_places=2):
            """Safely convert CharField value to float for Excel"""
            if not value:
                return 0.0
            try:
                # Remove commas and convert to Decimal first, then to float
                clean_value = str(value).replace(',', '').strip()
                decimal_value = Decimal(clean_value)
                return float(round(decimal_value, decimal_places))
            except (ValueError, InvalidOperation, TypeError):
                return 0.0

        # Write data rows
        for index, p in enumerate(portfolios, 1):
            # Calculate profit/loss safely
            try:
                cost_val = Decimal(str(p.cost_value).replace(',', '')) if p.cost_value else Decimal('0')
                market_val = Decimal(str(p.market_value).replace(',', '')) if p.market_value else Decimal('0')
                profit_loss = float(market_val - cost_val)
            except (ValueError, InvalidOperation, TypeError):
                profit_loss = 0.0

            ws.append([
                index,
                p.entity_name or '',
                p.folio_no or '',
                p.isin or '',
                p.scheme_name or '',
                safe_float_conversion(p.cost_value, 2),
                safe_float_conversion(p.unit_balance, 3),
                p.nav_date.strftime('%Y-%m-%d') if p.nav_date else 'NA',
                safe_float_conversion(p.market_value, 2),
                profit_loss,
                p.updated_at.strftime('%Y-%m-%d') if p.updated_at else 'NA',
                p.updated_by.username if p.updated_by else 'NA'
            ])

        # Optional: Adjust column widths
        for i, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in col) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_length, 50)  # Cap at 50 chars

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=cams_summary_filtered.xlsx'
        wb.save(response)
        return response


class DownloadDetailView(View):
    def get(self, request):
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        if not month or not year:
            return HttpResponse('Month and year required', status=400)
        
        # Get additional filters (optional for detail view)
        entity_filter = request.GET.get('entity_filter', '')
        isin_filter = request.GET.get('isin_filter', '')
        scheme_filter = request.GET.get('scheme_filter', '')
        
        # Start with month/year filter
        portfolios = CamsPortfolio.objects.filter(month=month, year=year)
        
        # Apply additional filters if provided
        if entity_filter:
            portfolios = portfolios.filter(entity_name__icontains=entity_filter)
        if isin_filter:
            portfolios = portfolios.filter(isin__icontains=isin_filter)
        if scheme_filter:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
        
        # Order by entity name, then folio number
        portfolios = portfolios.select_related('updated_by').order_by('entity_name', 'folio_no')
        
        # Create CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cams_detail_{year}_{month}.csv"'
        
        writer = csv.writer(response)
        # Write header
        writer.writerow([
            'Folio No', 
            'Entity Name', 
            'ISIN', 
            'Scheme Name', 
            'Cost Value', 
            'Unit Balance',
            'NAV Date',
            'Market Value',
            'Profit/Loss',
            'Month',
            'Year',
            'Updated On',
            'Updated By'
        ])
        
        # Write data rows
        for p in portfolios:
            writer.writerow([
                p.folio_no,
                p.entity_name,
                p.isin,
                p.scheme_name,
                f"{p.cost_value:.2f}",
                f"{p.unit_balance:.3f}",
                p.nav_date.strftime('%Y-%m-%d'),
                f"{p.market_value:.2f}",
                f"{p.profit_loss:.2f}",
                p.month,
                p.year,
                p.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                p.updated_by.username if p.updated_by else 'System'
            ])
        
        return response
    

@csrf_exempt  # remove this in production and use proper CSRF tokens
@require_POST
@login_required
def parse_cams_view(request):
        
        # # Delete all records from CAMSTransaction
        # CAMSTransaction.objects.all().delete()

        # # Delete all records from CAMSSummary
        # CAMSSummary.objects.all().delete()
        

    # try:
        process_cams_emails()
        
        import_cams_portfolio_with_pandas()
        result = process_cams_files("pdf_extracted_data")
        # return JsonResponse(result, safe=False)
        message = "Import ran successfully"
        return JsonResponse({"status": "success", "message": message})
   
    # except Exception as e:
    #     return JsonResponse({"status": "error", "message": str(e)}, status=500)
    






# # views.py
# from django.shortcuts import render
# from django.http import JsonResponse
# from django.core.paginator import Paginator
# from django.db.models import Q
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
# from django.views import View
# from .models import CamsPortfolio
# import json
# from datetime import datetime

# class CamsCarbyView(View):
#     """Main view for CAMS Carvy dashboard"""
    
#     def get(self, request):
#         context = {
#             'title': 'CAMS Carvy Dashboard',
#         }
#         return render(request, 'cams_carvy.html', context)

# @method_decorator(csrf_exempt, name='dispatch')
# class CamsPortfolioAPIView(View):
#     """API view for CAMS portfolio data"""
    
#     def get(self, request):
#         try:
#             # Get query parameters
#             draw = int(request.GET.get('draw', 1))
#             start = int(request.GET.get('start', 0))
#             length = int(request.GET.get('length', 10))
#             search_value = request.GET.get('search[value]', '')
            
#             # Column filters
#             entity_filter = request.GET.get('entity_name', '')
#             isin_filter = request.GET.get('isin', '')
#             scheme_filter = request.GET.get('scheme_name', '')
#             month_filter = request.GET.get('month', '')
            
#             # Build queryset
#             queryset = CamsPortfolio.objects.select_related('updated_by')
            
#             # Apply filters
#             if entity_filter:
#                 queryset = queryset.filter(entity_name__icontains=entity_filter)
            
#             if isin_filter:
#                 queryset = queryset.filter(isin__icontains=isin_filter)
            
#             if scheme_filter:
#                 queryset = queryset.filter(scheme_name__icontains=scheme_filter)
            
#             if month_filter:
#                 month_names = {
#                     'January': 1, 'February': 2, 'March': 3, 'April': 4,
#                     'May': 5, 'June': 6, 'July': 7, 'August': 8,
#                     'September': 9, 'October': 10, 'November': 11, 'December': 12
#                 }
#                 if month_filter in month_names:
#                     queryset = queryset.filter(month=month_names[month_filter])
            
#             # Global search
#             if search_value:
#                 queryset = queryset.filter(
#                     Q(folio_no__icontains=search_value) |
#                     Q(entity_name__icontains=search_value) |
#                     Q(isin__icontains=search_value) |
#                     Q(scheme_name__icontains=search_value)
#                 )
            
#             # Get total count before pagination
#             total_records = queryset.count()
            
#             # Apply pagination
#             paginator = Paginator(queryset, length)
#             page_number = (start // length) + 1
#             page_obj = paginator.get_page(page_number)
            
#             # Prepare data
#             data = []
#             for portfolio in page_obj:
#                 # Determine profit/loss styling
#                 market_value_display = f"₹{portfolio.market_value:,.2f}"
#                 if portfolio.is_profit:
#                     market_value_display += ' <i class="bi bi-arrow-up-right text-success"></i>'
#                     market_value_class = "text-success fw-bold"
#                 else:
#                     market_value_display += ' <i class="bi bi-arrow-down-right text-danger"></i>'
#                     market_value_class = "text-danger fw-bold"
                
#                 data.append({
#                     'folio_no': f'<a href="#" class="fw-bold text-primary">{portfolio.folio_no}</a>',
#                     'entity_name': portfolio.entity_name,
#                     'isin': portfolio.isin,
#                     'scheme_name': portfolio.scheme_name,
#                     'cost_value': f"₹{portfolio.cost_value:,.2f}",
#                     'unit_balance': f"{portfolio.unit_balance:,.3f}",
#                     'nav_date': portfolio.nav_date.strftime('%Y-%m-%d'),
#                     'market_value': f'<span class="{market_value_class}">{market_value_display}</span>',
#                     'updated_on': portfolio.updated_at.strftime('%Y-%m-%d'),
#                     'updated_by': portfolio.updated_by.get_full_name() or portfolio.updated_by.username,
#                 })
            
#             response = {
#                 'draw': draw,
#                 'recordsTotal': CamsPortfolio.objects.count(),
#                 'recordsFiltered': total_records,
#                 'data': data
#             }
            
#             return JsonResponse(response)
            
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class CamsSummaryAPIView(View):
#     """API view for CAMS summary statistics"""
    
#     def get(self, request):
#         try:
#             # Get filters
#             entity_filter = request.GET.get('entity_name', '')
#             isin_filter = request.GET.get('isin', '')
#             scheme_filter = request.GET.get('scheme_name', '')
#             month_filter = request.GET.get('month', '')
            
#             # Build queryset with filters
#             queryset = CamsPortfolio.objects.all()
            
#             if entity_filter:
#                 queryset = queryset.filter(entity_name__icontains=entity_filter)
            
#             if isin_filter:
#                 queryset = queryset.filter(isin__icontains=isin_filter)
            
#             if scheme_filter:
#                 queryset = queryset.filter(scheme_name__icontains=scheme_filter)
            
#             if month_filter:
#                 month_names = {
#                     'January': 1, 'February': 2, 'March': 3, 'April': 4,
#                     'May': 5, 'June': 6, 'July': 7, 'August': 8,
#                     'September': 9, 'October': 10, 'November': 11, 'December': 12
#                 }
#                 if month_filter in month_names:
#                     queryset = queryset.filter(month=month_names[month_filter])
            
#             # Get summary stats
#             stats = CamsPortfolio.get_summary_stats(queryset)
            
#             # Format response
#             response = {
#                 'total_entities': stats['total_entities'],
#                 'total_folios': stats['total_folios'],
#                 'total_market_value': f"₹{stats['total_market_value']:,.2f}" if stats['total_market_value'] else "₹0",
#                 'total_cost_value': f"₹{stats['total_cost_value']:,.2f}" if stats['total_cost_value'] else "₹0",
#             }
            
#             return JsonResponse(response)
            
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)







# views.py
import io
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
import openpyxl
from .models import CAMSTransaction, CAMSSummary



def download_current_month(request):
    today = timezone.now().date()
    start_date = today.replace(day=1)
    # Include the entire current month - end of today
    end_date = today
    filename = f"CAMS_Current_Month_{today.strftime('%Y_%m')}.xlsx"
    
    # Debug: Check if we have any data in current month
    print(f"Filtering from {start_date} to {end_date}")
    count = CAMSTransaction.objects.filter(transaction_date__range=[start_date, end_date]).count()
    print(f"Found {count} transactions in current month")
    
    return generate_excel(start_date, end_date, filename)

def download_previous_month(request):
    today = timezone.now().date()
    first_day_current = today.replace(day=1)
    last_day_previous = first_day_current - timedelta(days=1)
    start_date = last_day_previous.replace(day=1)
    end_date = last_day_previous
    filename = f"CAMS_Previous_Month_{last_day_previous.strftime('%Y_%m')}.xlsx"
    
    return generate_excel(start_date, end_date, filename)

def generate_excel(start_date, end_date, filename):
    # Get filtered transaction data based on date range
    transactions = CAMSTransaction.objects.filter(
        transaction_date__gte=start_date,
        transaction_date__lte=end_date
    ).order_by('entity_name', 'folio_number', 'transaction_date')
    
    # Get all summaries (current portfolio state)
    summaries = CAMSSummary.objects.all().order_by('entity_name', 'folio_number')
    
    print(f"Found {transactions.count()} transactions for period {start_date} to {end_date}")
    print(f"Found {summaries.count()} portfolio summaries")
    
    # Create Excel with both sheets
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    
    # Sheet 1: Transactions
    trans_sheet = workbook.create_sheet(title="Transactions")
    trans_headers = ['Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
                    'Transaction Date', 'Transaction Description', 'Value', 'Stamp Duty',
                    'Units', 'NAV', 'Unit Balance']
    
    for col, header in enumerate(trans_headers, 1):
        trans_sheet.cell(row=1, column=col, value=header)
    
    for row, transaction in enumerate(transactions, 2):
        trans_sheet.cell(row=row, column=1, value=transaction.entity_name)
        trans_sheet.cell(row=row, column=2, value=transaction.amc_name)
        trans_sheet.cell(row=row, column=3, value=transaction.folio_number)
        trans_sheet.cell(row=row, column=4, value=transaction.isin)
        trans_sheet.cell(row=row, column=5, value=transaction.scheme_name)
        trans_sheet.cell(row=row, column=6, value=transaction.transaction_date)
        trans_sheet.cell(row=row, column=7, value=transaction.transaction_description)
        trans_sheet.cell(row=row, column=8, value=float(transaction.value) if transaction.value else None)
        trans_sheet.cell(row=row, column=9, value=float(transaction.stamp_duty) if transaction.stamp_duty else None)
        trans_sheet.cell(row=row, column=10, value=float(transaction.units) if transaction.units else None)
        trans_sheet.cell(row=row, column=11, value=float(transaction.nav) if transaction.nav else None)
        trans_sheet.cell(row=row, column=12, value=float(transaction.unit_balance) if transaction.unit_balance else None)
    
    # Sheet 2: Portfolio Summary
    summary_sheet = workbook.create_sheet(title="Portfolio Summary")
    summary_headers = ['Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
                      'Closing Unit Balance', 'NAV', 'Total Cost Value', 'Market Value']
    
    for col, header in enumerate(summary_headers, 1):
        summary_sheet.cell(row=1, column=col, value=header)
    
    for row, summary in enumerate(summaries, 2):
        summary_sheet.cell(row=row, column=1, value=summary.entity_name)
        summary_sheet.cell(row=row, column=2, value=summary.amc_name)
        summary_sheet.cell(row=row, column=3, value=summary.folio_number)
        summary_sheet.cell(row=row, column=4, value=summary.isin)
        summary_sheet.cell(row=row, column=5, value=summary.scheme_name)
        summary_sheet.cell(row=row, column=6, value=float(summary.closing_unit_balance) if summary.closing_unit_balance else None)
        summary_sheet.cell(row=row, column=7, value=float(summary.nav) if summary.nav else None)
        summary_sheet.cell(row=row, column=8, value=float(summary.total_cost_value) if summary.total_cost_value else None)
        summary_sheet.cell(row=row, column=9, value=float(summary.market_value) if summary.market_value else None)
    
    # Save and return
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

