# views.py
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from .models import CamsPortfolio,CAMSTransaction, CAMSSummary
import json
from django.db.models import Sum, Count, Q
from django.shortcuts import get_object_or_404
import csv
from django.views import View
from .services import process_cams_files,import_cams_portfolio_with_pandas,process_cams_emails
from django.views.decorators.http import require_GET,require_POST
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.views import View
from django.db.models import Q
from .models import CamsPortfolio  # Adjust import as needed

from django.views import View
from django.http import HttpResponse
from django.db.models import Q
import openpyxl
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation
from .models import CamsPortfolio  # Replace 'your_app' with your actual app name

from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q, Count
from decimal import Decimal, InvalidOperation

@login_required
def cams_carvy_dashboard(request):
    """Main dashboard view"""
    # Get all portfolios for summary stats
    all_portfolios = CamsPortfolio.objects.all()
    summary_stats = CamsPortfolio.get_summary_stats(all_portfolios)
    
    context = {
        'summary_stats': summary_stats,
    }
    return render(request, 'test.html', context)

# @login_required
# def get_portfolio_data(request):
#     """Ajax endpoint for portfolio data with filtering, pagination, and search"""
#     # Get query parameters
#     draw = int(request.GET.get('draw', 1))
#     start = int(request.GET.get('start', 0))
#     length = int(request.GET.get('length', 10))
#     search_value = request.GET.get('search[value]', '')
    
#     # Column search parameters
#     entity_filter = request.GET.get('entity_filter', '')
#     isin_filter = request.GET.get('isin_filter', '')
#     scheme_filter = request.GET.get('scheme_filter', '')
#     month_filter = request.GET.get('month_filter', '')
    
#     # Individual column searches
#     folio_search = request.GET.get('col_0_search', '')
#     entity_search = request.GET.get('col_1_search', '')
#     isin_search = request.GET.get('col_2_search', '')
#     scheme_search = request.GET.get('col_3_search', '')
#     cost_search = request.GET.get('col_4_search', '')
#     unit_search = request.GET.get('col_5_search', '')
#     nav_search = request.GET.get('col_6_search', '')
#     market_search = request.GET.get('col_7_search', '')
#     updated_search = request.GET.get('col_8_search', '')
#     user_search = request.GET.get('col_9_search', '')
    
#     # Start with all portfolios
#     portfolios = CamsPortfolio.objects.all()
    
#     # Apply filters
#     if entity_filter:
#         portfolios = portfolios.filter(entity_name__icontains=entity_filter)
#     if isin_filter:
#         portfolios = portfolios.filter(isin__icontains=isin_filter)
#     if scheme_filter:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
#     if month_filter:
#         portfolios = portfolios.filter(month=month_filter)
    
#     # Apply global search
#     if search_value:
#         portfolios = portfolios.filter(
#             Q(folio_no__icontains=search_value) |
#             Q(entity_name__icontains=search_value) |
#             Q(isin__icontains=search_value) |
#             Q(scheme_name__icontains=search_value)
#         )
    
#     # Apply column-specific searches
#     if folio_search:
#         portfolios = portfolios.filter(folio_no__icontains=folio_search)
#     if entity_search:
#         portfolios = portfolios.filter(entity_name__icontains=entity_search)
#     if isin_search:
#         portfolios = portfolios.filter(isin__icontains=isin_search)
#     if scheme_search:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
#     if cost_search:
#         portfolios = portfolios.filter(cost_value__icontains=cost_search)
#     if unit_search:
#         portfolios = portfolios.filter(unit_balance__icontains=unit_search)
#     if nav_search:
#         portfolios = portfolios.filter(nav_date__icontains=nav_search)
#     if market_search:
#         portfolios = portfolios.filter(market_value__icontains=market_search)
#     if updated_search:
#         portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
#     if user_search:
#         portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
#     # Get total count before pagination
#     total_records = portfolios.count()
    
#     # Apply ordering (default by updated_at desc)
#     order_column = request.GET.get('order[0][column]', '8')  # Default to updated_at
#     order_dir = request.GET.get('order[0][dir]', 'desc')
    
#     order_columns = {
#         '0': 'folio_no',
#         '1': 'entity_name', 
#         '2': 'isin',
#         '3': 'scheme_name',
#         '4': 'cost_value',
#         '5': 'unit_balance',
#         '6': 'nav_date',
#         '7': 'market_value',
#         '8': 'updated_at',
#         '9': 'updated_by__username'
#     }
    
#     order_field = order_columns.get(order_column, 'updated_at')
#     if order_dir == 'desc':
#         order_field = f'-{order_field}'
    
#     portfolios = portfolios.select_related('updated_by').order_by(order_field)
    
#     # Apply pagination
#     portfolios = portfolios[start:start + length]
    
#     # Prepare data for DataTables
#     data = []
#     for portfolio in portfolios:
#         profit_loss = portfolio.profit_loss
#         is_profit = profit_loss >= 0
#         profit_class = 'text-success' if is_profit else 'text-danger'
#         profit_icon = 'bi-arrow-up-right' if is_profit else 'bi-arrow-down-right'
        
#         data.append([
#             f'<a href="#" class="fw-bold text-primary">{portfolio.folio_no}</a>',
#             portfolio.entity_name,
#             f'<span class="editable-isin" data-id="{portfolio.id}" data-value="{portfolio.isin}">{portfolio.isin}</span>',
#             portfolio.scheme_name,
#             f'₹{portfolio.cost_value:,.2f}',
#             f'{portfolio.unit_balance:,.3f}',
#             portfolio.nav_date.strftime('%Y-%m-%d'),
#             # f'<span class="{profit_class} fw-bold">₹{portfolio.market_value:,.2f} <i class="bi {profit_icon}"></i></span>',
#             # portfolio.updated_at.strftime('%Y-%m-%d'),
#             # portfolio.updated_by.username if portfolio.updated_by else 'System'
#             f'<span class="{profit_class} fw-bold">₹{portfolio.market_value:,.2f} <i class="bi {profit_icon}"></i></span>',
#             portfolio.updated_at.strftime('%Y-%m-%d') if portfolio.updated_at else 'NA',
#             portfolio.updated_by.username if portfolio.updated_by else 'NA',

#         ])
    
#     return JsonResponse({
#         'draw': draw,
#         'recordsTotal': CamsPortfolio.objects.count(),
#         'recordsFiltered': total_records,
#         'data': data
#     })



# @csrf_exempt
# @login_required
# def update_isin(request):
#     """Ajax endpoint to update ISIN"""
#     if request.method == 'POST':
#         try:
#             data = json.loads(request.body)
#             portfolio_id = data.get('id')
#             new_isin = data.get('isin', '').strip()
            
#             if not new_isin:
#                 return JsonResponse({'success': False, 'error': 'ISIN cannot be empty'})
            
#             if len(new_isin) != 12:
#                 return JsonResponse({'success': False, 'error': 'ISIN must be 12 characters long'})
            
#             portfolio = get_object_or_404(CamsPortfolio, id=portfolio_id)
#             portfolio.isin = new_isin
#             portfolio.updated_by = request.user
#             portfolio.save()
            
#             return JsonResponse({'success': True, 'isin': new_isin})
            
#         except Exception as e:
#             return JsonResponse({'success': False, 'error': str(e)})
    
#     return JsonResponse({'success': False, 'error': 'Invalid request method'})

def format_indian_number(number, decimals=2):
    """
    Convert a number to Indian style format with commas.
    Example: 490028384.45 -> 49,00,28,384.45
    """
    if number is None:
        return f"0.{ '0'*decimals }"

    try:
        number = float(number)
    except (ValueError, TypeError):
        return str(number)

    # Split integer and decimal parts
    int_part, dec_part = divmod(number, 1)
    dec_part = round(dec_part * (10 ** decimals))  # handle decimals

    int_part_str = str(int(int_part))

    if len(int_part_str) <= 3:
        formatted_int = int_part_str
    else:
        last3 = int_part_str[-3:]
        rest = int_part_str[:-3]
        rest_groups = []
        while len(rest) > 2:
            rest_groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            rest_groups.insert(0, rest)
        formatted_int = ','.join(rest_groups) + ',' + last3

    return f"{formatted_int}.{dec_part:0{decimals}d}"





@login_required
def get_portfolio_data(request):
    """Ajax endpoint for portfolio data with filtering, pagination, and search"""
    # Get query parameters
    draw = int(request.GET.get('draw', 1))
    start = int(request.GET.get('start', 0))
    length = int(request.GET.get('length', 10))
    search_value = request.GET.get('search[value]', '')
    
    # Column search parameters
    entity_filter = request.GET.get('entity_filter', '')
    isin_filter = request.GET.get('isin_filter', '')
    scheme_filter = request.GET.get('scheme_filter', '')
    month_filter = request.GET.get('month_filter', '')
    
    # Individual column searches - UPDATED MAPPING
    sno_search = request.GET.get('col_0_search', '')        # S.No (not used for filtering)
    entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped position)
    folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped position)
    isin_search = request.GET.get('col_3_search', '')       # ISIN
    scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
    cost_search = request.GET.get('col_5_search', '')       # Cost Value
    unit_search = request.GET.get('col_6_search', '')       # Unit Balance
    nav_search = request.GET.get('col_7_search', '')        # NAV Date
    market_search = request.GET.get('col_8_search', '')     # Market Value
    updated_search = request.GET.get('col_9_search', '')    # Updated On
    user_search = request.GET.get('col_10_search', '')      # Updated By
    
    # Start with all portfolios
    portfolios = CamsPortfolio.objects.all()
    
    # Apply filters
    if entity_filter:
        portfolios = portfolios.filter(entity_name__icontains=entity_filter)
    if isin_filter:
        portfolios = portfolios.filter(isin__icontains=isin_filter)
    if scheme_filter:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
    if month_filter:
        portfolios = portfolios.filter(month=month_filter)
    
    # Apply global search
    if search_value:
        portfolios = portfolios.filter(
            Q(folio_no__icontains=search_value) |
            Q(entity_name__icontains=search_value) |
            Q(isin__icontains=search_value) |
            Q(scheme_name__icontains=search_value)
        )
    
    # Apply column-specific searches
    if folio_search:
        portfolios = portfolios.filter(folio_no__icontains=folio_search)
    if entity_search:
        portfolios = portfolios.filter(entity_name__icontains=entity_search)
    if isin_search:
        portfolios = portfolios.filter(isin__icontains=isin_search)
    if scheme_search:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
    if cost_search:
        portfolios = portfolios.filter(cost_value__icontains=cost_search)
    if unit_search:
        portfolios = portfolios.filter(unit_balance__icontains=unit_search)
    if nav_search:
        portfolios = portfolios.filter(nav_date__icontains=nav_search)
    if market_search:
        portfolios = portfolios.filter(market_value__icontains=market_search)
    if updated_search:
        portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
    if user_search:
        portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
    # Get total count before pagination
    total_records = portfolios.count()
    
    # Apply ordering - UPDATED COLUMN MAPPING
    order_column = request.GET.get('order[0][column]', '9')  # Default to updated_at (column 9)
    order_dir = request.GET.get('order[0][dir]', 'desc')
    
    order_columns = {
        '0': 'id',                      # S.No - use ID for ordering
        '1': 'entity_name',             # Entity Name (swapped)
        '2': 'folio_no',                # Folio No (swapped)
        '3': 'isin',                    # ISIN
        '4': 'scheme_name',             # Scheme Name
        '5': 'cost_value',              # Cost Value
        '6': 'unit_balance',            # Unit Balance
        '7': 'nav_date',                # NAV Date
        '8': 'market_value',            # Market Value
        '9': 'updated_at',              # Updated On
        '10': 'updated_by__username'    # Updated By
    }
    
    order_field = order_columns.get(order_column, 'updated_at')
    if order_dir == 'desc':
        order_field = f'-{order_field}'
    
    portfolios = portfolios.select_related('updated_by').order_by(order_field)
    
    # Apply pagination
    portfolios = portfolios[start:start + length]
    
    # Prepare data for DataTables with S.No - UPDATED DATA ORDER
    data = []
    for index, portfolio in enumerate(portfolios):
        # print(type(portfolio.market_value),"'////")
        # if ',' in portfolio.market_value :
        #         market_value =  portfolio.market_value.replace(',','')

        profit_loss = portfolio.profit_loss
        is_profit = profit_loss >= 0
        profit_class = 'text-success' if is_profit else 'text-danger'
        profit_icon = 'bi-arrow-up-right' if is_profit else 'bi-arrow-down-right'
        
        # Calculate S.No based on pagination
        sno = start + index + 1
        
        data.append([
            sno,  # Column 0: S.No
            portfolio.entity_name,  # Column 1: Entity Name (swapped)
            f'<span href="#" class=" ">{portfolio.folio_no}</span>',  # Column 2: Folio No (swapped)
            f'<span class="editable-isin" data-id="{portfolio.id}" data-value="{portfolio.isin}">{portfolio.isin}</span>',  # Column 3: ISIN
            portfolio.scheme_name,  # Column 4: Scheme Name
            
            # f'₹{portfolio.cost_value:,.2f}',  # Column 5: Cost Value
            f'₹{format_indian_number(str(portfolio.cost_value).replace(",",""))}',  # Column 5: Cost Value
            # f'₹{portfolio.unit_balance:,.2f}',  # Column 6: Unit Balance
            f'₹{format_indian_number(str(portfolio.unit_balance).replace(",",""))}',  # Column 6: Unit Balance

            portfolio.nav_date.strftime('%d-%m-%Y'),  # Column 7: NAV Date
            
            # f'<span class="{profit_class} ">₹{portfolio.market_value:,.2f}',  # Column 8: Market Value
            f'<span  ">₹{format_indian_number(str(portfolio.market_value))}',  # Column 8: Market Value
            
            
            portfolio.updated_at.strftime('%d-%m-%Y') if portfolio.updated_at else 'NA',  # Column 9: Updated On
            portfolio.updated_by.username if portfolio.updated_by else 'NA',  # Column 10: Updated By
        ])
    
    return JsonResponse({
        'draw': draw,
        'recordsTotal': CamsPortfolio.objects.count(),
        'recordsFiltered': total_records,
        'data': data
    })


# @login_required
# def get_summary_stats(request):
#     """Ajax endpoint for summary statistics with filtering"""
    
#     # Get all the same filter parameters as get_portfolio_data - UPDATED MAPPING
#     entity_filter = request.GET.get('entity_filter', '')
#     isin_filter = request.GET.get('isin_filter', '')
#     scheme_filter = request.GET.get('scheme_filter', '')
#     month_filter = request.GET.get('month_filter', '')
    
#     # Global search
#     global_search = request.GET.get('global_search', '')
    
#     # Individual column searches - UPDATED MAPPING
#     sno_search = request.GET.get('col_0_search', '')        # S.No (not used)
#     entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped)
#     folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped)
#     isin_search = request.GET.get('col_3_search', '')       # ISIN
#     scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
#     cost_search = request.GET.get('col_5_search', '')       # Cost Value
#     unit_search = request.GET.get('col_6_search', '')       # Unit Balance
#     nav_search = request.GET.get('col_7_search', '')        # NAV Date
#     market_search = request.GET.get('col_8_search', '')     # Market Value
#     updated_search = request.GET.get('col_9_search', '')    # Updated On
#     user_search = request.GET.get('col_10_search', '')      # Updated By
    
#     # Start with all portfolios - SAME FILTERING LOGIC AS get_portfolio_data
#     portfolios = CamsPortfolio.objects.all()
    
#     # Apply main filters
#     if entity_filter:
#         portfolios = portfolios.filter(entity_name__icontains=entity_filter)
#     if isin_filter:
#         portfolios = portfolios.filter(isin__icontains=isin_filter)
#     if scheme_filter:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
#     if month_filter:
#         portfolios = portfolios.filter(month=month_filter)
    
#     # Apply global search
#     if global_search:
#         portfolios = portfolios.filter(
#             Q(folio_no__icontains=global_search) |
#             Q(entity_name__icontains=global_search) |
#             Q(isin__icontains=global_search) |
#             Q(scheme_name__icontains=global_search)
#         )
    
#     # Apply column-specific searches
#     if folio_search:
#         portfolios = portfolios.filter(folio_no__icontains=folio_search)
#     if entity_search:
#         portfolios = portfolios.filter(entity_name__icontains=entity_search)
#     if isin_search:
#         portfolios = portfolios.filter(isin__icontains=isin_search)
#     if scheme_search:
#         portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
#     if cost_search:
#         portfolios = portfolios.filter(cost_value__icontains=cost_search)
#     if unit_search:
#         portfolios = portfolios.filter(unit_balance__icontains=unit_search)
#     if nav_search:
#         portfolios = portfolios.filter(nav_date__icontains=nav_search)
#     if market_search:
#         portfolios = portfolios.filter(market_value__icontains=market_search)
#     if updated_search:
#         portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
#     if user_search:
#         portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
#     # Calculate statistics on the filtered queryset
#     stats = portfolios.aggregate(
#         total_market_value=Sum('market_value'),
#         total_cost_value=Sum('cost_value'),
#         total_folios=Count('id'),
#         total_entities=Count('entity_name', distinct=True)
#     )
    
#     return JsonResponse({
#         'total_entities': stats['total_entities'] or 0,
#         'total_folios': stats['total_folios'] or 0,
#         'total_market_value': float(stats['total_market_value'] or 0),
#         'total_cost_value': float(stats['total_cost_value'] or 0)
#     })





@login_required
def get_summary_stats(request):
    """Ajax endpoint for summary statistics with filtering"""
    
    # Get all the same filter parameters as get_portfolio_data - UPDATED MAPPING
    entity_filter = request.GET.get('entity_filter', '')
    isin_filter = request.GET.get('isin_filter', '')
    scheme_filter = request.GET.get('scheme_filter', '')
    month_filter = request.GET.get('month_filter', '')
    
    # Global search
    global_search = request.GET.get('global_search', '')
    
    # Individual column searches - UPDATED MAPPING
    sno_search = request.GET.get('col_0_search', '')        # S.No (not used)
    entity_search = request.GET.get('col_1_search', '')     # Entity Name (swapped)
    folio_search = request.GET.get('col_2_search', '')      # Folio No (swapped)
    isin_search = request.GET.get('col_3_search', '')       # ISIN
    scheme_search = request.GET.get('col_4_search', '')     # Scheme Name
    cost_search = request.GET.get('col_5_search', '')       # Cost Value
    unit_search = request.GET.get('col_6_search', '')       # Unit Balance
    nav_search = request.GET.get('col_7_search', '')        # NAV Date
    market_search = request.GET.get('col_8_search', '')     # Market Value
    updated_search = request.GET.get('col_9_search', '')    # Updated On
    user_search = request.GET.get('col_10_search', '')      # Updated By
    
    # Start with all portfolios - SAME FILTERING LOGIC AS get_portfolio_data
    portfolios = CamsPortfolio.objects.all()
    
    # Apply main filters
    if entity_filter:
        portfolios = portfolios.filter(entity_name__icontains=entity_filter)
    if isin_filter:
        portfolios = portfolios.filter(isin__icontains=isin_filter)
    if scheme_filter:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
    if month_filter:
        portfolios = portfolios.filter(month=month_filter)
    
    # Apply global search
    if global_search:
        portfolios = portfolios.filter(
            Q(folio_no__icontains=global_search) |
            Q(entity_name__icontains=global_search) |
            Q(isin__icontains=global_search) |
            Q(scheme_name__icontains=global_search)
        )
    
    # Apply column-specific searches
    if folio_search:
        portfolios = portfolios.filter(folio_no__icontains=folio_search)
    if entity_search:
        portfolios = portfolios.filter(entity_name__icontains=entity_search)
    if isin_search:
        portfolios = portfolios.filter(isin__icontains=isin_search)
    if scheme_search:
        portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
    if cost_search:
        portfolios = portfolios.filter(cost_value__icontains=cost_search)
    if unit_search:
        portfolios = portfolios.filter(unit_balance__icontains=unit_search)
    if nav_search:
        portfolios = portfolios.filter(nav_date__icontains=nav_search)
    if market_search:
        portfolios = portfolios.filter(market_value__icontains=market_search)
    if updated_search:
        portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
    if user_search:
        portfolios = portfolios.filter(updated_by__username__icontains=user_search)
    
    # Calculate statistics manually since CharField can't use Sum()
    total_market_value = Decimal('0')
    total_cost_value = Decimal('0')
    
    for portfolio in portfolios:
        try:
            if portfolio.market_value:
                market_val = Decimal(str(portfolio.market_value).replace(',', ''))
                total_market_value += market_val
        except (ValueError, InvalidOperation, TypeError):
            pass
            
        try:
            if portfolio.cost_value:
                cost_val = Decimal(str(portfolio.cost_value).replace(',', ''))
                total_cost_value += cost_val
        except (ValueError, InvalidOperation, TypeError):
            pass
    
    # Get count statistics using Django ORM
    count_stats = portfolios.aggregate(
        total_folios=Count('id'),
        total_entities=Count('entity_name', distinct=True)
    )
    
    return JsonResponse({
        'total_entities': count_stats['total_entities'] or 0,
        'total_folios': count_stats['total_folios'] or 0,
        'total_market_value': float(total_market_value),
        'total_cost_value': float(total_cost_value)
    })


@csrf_exempt
@login_required
def update_isin(request):
    """Ajax endpoint to update ISIN with updated validation (8-20 characters)"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            portfolio_id = data.get('id')
            new_isin = data.get('isin', '').strip()
            
            if not new_isin:
                return JsonResponse({'success': False, 'error': 'ISIN cannot be empty'})
            
            # UPDATED: Changed validation from 12 chars to 8-20 chars
            if len(new_isin) < 8 or len(new_isin) > 20:
                return JsonResponse({'success': False, 'error': 'ISIN must be 8 to 20 characters long'})
            
            portfolio = get_object_or_404(CamsPortfolio, id=portfolio_id)
            portfolio.isin = new_isin
            portfolio.updated_by = request.user
            portfolio.save()
            
            return JsonResponse({'success': True, 'isin': new_isin})
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})





class DownloadSummaryView(View):
    def get(self, request):
        # Filters
        entity_filter = request.GET.get('entity_filter', '')
        isin_filter = request.GET.get('isin_filter', '')
        scheme_filter = request.GET.get('scheme_filter', '')
        month_filter = request.GET.get('month_filter', '')
        global_search = request.GET.get('global_search', '')

        # Column-specific searches
        sno_search = request.GET.get('col_0_search', '')
        entity_search = request.GET.get('col_1_search', '')
        folio_search = request.GET.get('col_2_search', '')
        isin_search = request.GET.get('col_3_search', '')
        scheme_search = request.GET.get('col_4_search', '')
        cost_search = request.GET.get('col_5_search', '')
        unit_search = request.GET.get('col_6_search', '')
        nav_search = request.GET.get('col_7_search', '')
        market_search = request.GET.get('col_8_search', '')
        updated_search = request.GET.get('col_9_search', '')
        user_search = request.GET.get('col_10_search', '')

        # Queryset
        portfolios = CamsPortfolio.objects.all()

        if entity_filter:
            portfolios = portfolios.filter(entity_name__icontains=entity_filter)
        if isin_filter:
            portfolios = portfolios.filter(isin__icontains=isin_filter)
        if scheme_filter:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
        if month_filter:
            portfolios = portfolios.filter(month=month_filter)

        if global_search:
            portfolios = portfolios.filter(
                Q(folio_no__icontains=global_search) |
                Q(entity_name__icontains=global_search) |
                Q(isin__icontains=global_search) |
                Q(scheme_name__icontains=global_search)
            )

        # Column-specific filtering
        if folio_search:
            portfolios = portfolios.filter(folio_no__icontains=folio_search)
        if entity_search:
            portfolios = portfolios.filter(entity_name__icontains=entity_search)
        if isin_search:
            portfolios = portfolios.filter(isin__icontains=isin_search)
        if scheme_search:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_search)
        if cost_search:
            portfolios = portfolios.filter(cost_value__icontains=cost_search)
        if unit_search:
            portfolios = portfolios.filter(unit_balance__icontains=unit_search)
        if nav_search:
            portfolios = portfolios.filter(nav_date__icontains=nav_search)
        if market_search:
            portfolios = portfolios.filter(market_value__icontains=market_search)
        if updated_search:
            portfolios = portfolios.filter(updated_at__date__icontains=updated_search)
        if user_search:
            portfolios = portfolios.filter(updated_by__username__icontains=user_search)

        portfolios = portfolios.select_related('updated_by').order_by('-updated_at')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CAMS Summary"

        # Header
        headers = [
            'S.No',
            'Entity Name',
            'Folio No',
            'ISIN',
            'Scheme Name',
            'Cost Value',
            'Unit Balance',
            'NAV Date',
            'Market Value',
            'Profit/Loss',
            'Updated On',
            'Updated By'
        ]
        ws.append(headers)

        # Helper function to safely convert string to float
        def safe_float_conversion(value, decimal_places=2):
            """Safely convert CharField value to float for Excel"""
            if not value:
                return 0.0
            try:
                # Remove commas and convert to Decimal first, then to float
                clean_value = str(value).replace(',', '').strip()
                decimal_value = Decimal(clean_value)
                return float(round(decimal_value, decimal_places))
            except (ValueError, InvalidOperation, TypeError):
                return 0.0

        # Write data rows
        for index, p in enumerate(portfolios, 1):
            # Calculate profit/loss safely
            try:
                cost_val = Decimal(str(p.cost_value).replace(',', '')) if p.cost_value else Decimal('0')
                market_val = Decimal(str(p.market_value).replace(',', '')) if p.market_value else Decimal('0')
                profit_loss = float(market_val - cost_val)
            except (ValueError, InvalidOperation, TypeError):
                profit_loss = 0.0

            ws.append([
                index,
                p.entity_name or '',
                p.folio_no or '',
                p.isin or '',
                p.scheme_name or '',
                safe_float_conversion(p.cost_value, 2),
                safe_float_conversion(p.unit_balance, 3),
                p.nav_date.strftime('%Y-%m-%d') if p.nav_date else 'NA',
                safe_float_conversion(p.market_value, 2),
                profit_loss,
                p.updated_at.strftime('%Y-%m-%d') if p.updated_at else 'NA',
                p.updated_by.username if p.updated_by else 'NA'
            ])

        # Optional: Adjust column widths
        for i, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in col) + 2
            ws.column_dimensions[get_column_letter(i)].width = min(max_length, 50)  # Cap at 50 chars

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=cams_summary_filtered.xlsx'
        wb.save(response)
        return response


class DownloadDetailView(View):
    def get(self, request):
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        if not month or not year:
            return HttpResponse('Month and year required', status=400)
        
        # Get additional filters (optional for detail view)
        entity_filter = request.GET.get('entity_filter', '')
        isin_filter = request.GET.get('isin_filter', '')
        scheme_filter = request.GET.get('scheme_filter', '')
        
        # Start with month/year filter
        portfolios = CamsPortfolio.objects.filter(month=month, year=year)
        
        # Apply additional filters if provided
        if entity_filter:
            portfolios = portfolios.filter(entity_name__icontains=entity_filter)
        if isin_filter:
            portfolios = portfolios.filter(isin__icontains=isin_filter)
        if scheme_filter:
            portfolios = portfolios.filter(scheme_name__icontains=scheme_filter)
        
        # Order by entity name, then folio number
        portfolios = portfolios.select_related('updated_by').order_by('entity_name', 'folio_no')
        
        # Create CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="cams_detail_{year}_{month}.csv"'
        
        writer = csv.writer(response)
        # Write header
        writer.writerow([
            'Folio No', 
            'Entity Name', 
            'ISIN', 
            'Scheme Name', 
            'Cost Value', 
            'Unit Balance',
            'NAV Date',
            'Market Value',
            'Profit/Loss',
            'Month',
            'Year',
            'Updated On',
            'Updated By'
        ])
        
        # Write data rows
        for p in portfolios:
            writer.writerow([
                p.folio_no,
                p.entity_name,
                p.isin,
                p.scheme_name,
                f"{p.cost_value:.2f}",
                f"{p.unit_balance:.3f}",
                p.nav_date.strftime('%Y-%m-%d'),
                f"{p.market_value:.2f}",
                f"{p.profit_loss:.2f}",
                p.month,
                p.year,
                p.updated_at.strftime('%Y-%m-%d %H:%M:%S'),
                p.updated_by.username if p.updated_by else 'System'
            ])
        
        return response
    


# ================ delete current month data ======================
from django.utils import timezone
from datetime import timedelta
from .models import CAMSTransaction, CAMSSummary

def delete_current_month_data():
    today = timezone.now().date()

    # If today is 1st → current month = previous month
    if today.day == 1:
        first_day_current = today.replace(day=1)
        current_month_start = (first_day_current - timedelta(days=1)).replace(day=1)  # 1st of prev month
    else:
        current_month_start = today.replace(day=1)  # 1st of current month

    # Last day of current month (exclusive end for filter)
    if current_month_start.month == 12:
        next_month_start = current_month_start.replace(year=current_month_start.year + 1, month=1, day=1)
    else:
        next_month_start = current_month_start.replace(month=current_month_start.month + 1, day=1)

    # Debug log
    print(f"Deleting transactions from {current_month_start} to {next_month_start - timedelta(days=1)}")

    # Delete only current month’s transactions
    CAMSTransaction.objects.filter(
        transaction_date__gte=current_month_start,
        transaction_date__lt=next_month_start
    ).delete()

    # Same logic for summaries (optional, if you want)
    CAMSSummary.objects.filter(
        nav_date__gte=current_month_start,
        nav_date__lt=next_month_start
    ).delete()

# ================ delete current month data ======================



@csrf_exempt  # remove this in production and use proper CSRF tokens
@require_POST
@login_required
def parse_cams_view(request):
        

    # try:
        print(process_cams_emails())

        delete_current_month_data()
        
        import_cams_portfolio_with_pandas()

        result = process_cams_files("pdf_extracted_data")
        # return JsonResponse(result, safe=False)
        message = "Import ran successfully"
        return JsonResponse({"status": "success", "message": message})
   
    # except Exception as e:
    #     return JsonResponse({"status": "error", "message": str(e)}, status=500)
    






# # views.py
# from django.shortcuts import render
# from django.http import JsonResponse
# from django.core.paginator import Paginator
# from django.db.models import Q
# from django.views.decorators.csrf import csrf_exempt
# from django.utils.decorators import method_decorator
# from django.views import View
# from .models import CamsPortfolio
# import json
# from datetime import datetime

# class CamsCarbyView(View):
#     """Main view for CAMS Carvy dashboard"""
    
#     def get(self, request):
#         context = {
#             'title': 'CAMS Carvy Dashboard',
#         }
#         return render(request, 'cams_carvy.html', context)

# @method_decorator(csrf_exempt, name='dispatch')
# class CamsPortfolioAPIView(View):
#     """API view for CAMS portfolio data"""
    
#     def get(self, request):
#         try:
#             # Get query parameters
#             draw = int(request.GET.get('draw', 1))
#             start = int(request.GET.get('start', 0))
#             length = int(request.GET.get('length', 10))
#             search_value = request.GET.get('search[value]', '')
            
#             # Column filters
#             entity_filter = request.GET.get('entity_name', '')
#             isin_filter = request.GET.get('isin', '')
#             scheme_filter = request.GET.get('scheme_name', '')
#             month_filter = request.GET.get('month', '')
            
#             # Build queryset
#             queryset = CamsPortfolio.objects.select_related('updated_by')
            
#             # Apply filters
#             if entity_filter:
#                 queryset = queryset.filter(entity_name__icontains=entity_filter)
            
#             if isin_filter:
#                 queryset = queryset.filter(isin__icontains=isin_filter)
            
#             if scheme_filter:
#                 queryset = queryset.filter(scheme_name__icontains=scheme_filter)
            
#             if month_filter:
#                 month_names = {
#                     'January': 1, 'February': 2, 'March': 3, 'April': 4,
#                     'May': 5, 'June': 6, 'July': 7, 'August': 8,
#                     'September': 9, 'October': 10, 'November': 11, 'December': 12
#                 }
#                 if month_filter in month_names:
#                     queryset = queryset.filter(month=month_names[month_filter])
            
#             # Global search
#             if search_value:
#                 queryset = queryset.filter(
#                     Q(folio_no__icontains=search_value) |
#                     Q(entity_name__icontains=search_value) |
#                     Q(isin__icontains=search_value) |
#                     Q(scheme_name__icontains=search_value)
#                 )
            
#             # Get total count before pagination
#             total_records = queryset.count()
            
#             # Apply pagination
#             paginator = Paginator(queryset, length)
#             page_number = (start // length) + 1
#             page_obj = paginator.get_page(page_number)
            
#             # Prepare data
#             data = []
#             for portfolio in page_obj:
#                 # Determine profit/loss styling
#                 market_value_display = f"₹{portfolio.market_value:,.2f}"
#                 if portfolio.is_profit:
#                     market_value_display += ' <i class="bi bi-arrow-up-right text-success"></i>'
#                     market_value_class = "text-success fw-bold"
#                 else:
#                     market_value_display += ' <i class="bi bi-arrow-down-right text-danger"></i>'
#                     market_value_class = "text-danger fw-bold"
                
#                 data.append({
#                     'folio_no': f'<a href="#" class="fw-bold text-primary">{portfolio.folio_no}</a>',
#                     'entity_name': portfolio.entity_name,
#                     'isin': portfolio.isin,
#                     'scheme_name': portfolio.scheme_name,
#                     'cost_value': f"₹{portfolio.cost_value:,.2f}",
#                     'unit_balance': f"{portfolio.unit_balance:,.3f}",
#                     'nav_date': portfolio.nav_date.strftime('%Y-%m-%d'),
#                     'market_value': f'<span class="{market_value_class}">{market_value_display}</span>',
#                     'updated_on': portfolio.updated_at.strftime('%Y-%m-%d'),
#                     'updated_by': portfolio.updated_by.get_full_name() or portfolio.updated_by.username,
#                 })
            
#             response = {
#                 'draw': draw,
#                 'recordsTotal': CamsPortfolio.objects.count(),
#                 'recordsFiltered': total_records,
#                 'data': data
#             }
            
#             return JsonResponse(response)
            
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)

# class CamsSummaryAPIView(View):
#     """API view for CAMS summary statistics"""
    
#     def get(self, request):
#         try:
#             # Get filters
#             entity_filter = request.GET.get('entity_name', '')
#             isin_filter = request.GET.get('isin', '')
#             scheme_filter = request.GET.get('scheme_name', '')
#             month_filter = request.GET.get('month', '')
            
#             # Build queryset with filters
#             queryset = CamsPortfolio.objects.all()
            
#             if entity_filter:
#                 queryset = queryset.filter(entity_name__icontains=entity_filter)
            
#             if isin_filter:
#                 queryset = queryset.filter(isin__icontains=isin_filter)
            
#             if scheme_filter:
#                 queryset = queryset.filter(scheme_name__icontains=scheme_filter)
            
#             if month_filter:
#                 month_names = {
#                     'January': 1, 'February': 2, 'March': 3, 'April': 4,
#                     'May': 5, 'June': 6, 'July': 7, 'August': 8,
#                     'September': 9, 'October': 10, 'November': 11, 'December': 12
#                 }
#                 if month_filter in month_names:
#                     queryset = queryset.filter(month=month_names[month_filter])
            
#             # Get summary stats
#             stats = CamsPortfolio.get_summary_stats(queryset)
            
#             # Format response
#             response = {
#                 'total_entities': stats['total_entities'],
#                 'total_folios': stats['total_folios'],
#                 'total_market_value': f"₹{stats['total_market_value']:,.2f}" if stats['total_market_value'] else "₹0",
#                 'total_cost_value': f"₹{stats['total_cost_value']:,.2f}" if stats['total_cost_value'] else "₹0",
#             }
            
#             return JsonResponse(response)
            
#         except Exception as e:
#             return JsonResponse({'error': str(e)}, status=500)







# views.py
import io
from datetime import datetime, timedelta
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
import openpyxl
from .models import CAMSTransaction, CAMSSummary



# def download_current_month(request):
#     today = timezone.now().date()
#     start_date = today.replace(day=1)
#     # Include the entire current month - end of today
#     end_date = today
#     filename = f"CAMS_Current_Month_{today.strftime('%Y_%m')}.xlsx"
    
#     # Debug: Check if we have any data in current month
#     print(f"Filtering from {start_date} to {end_date}")
#     count = CAMSTransaction.objects.filter(transaction_date__range=[start_date, end_date]).count()
#     print(f"Found {count} transactions in current month")
    
#     return generate_excel(start_date, end_date, filename)

# import io
# from datetime import timedelta, date, datetime
# from django.http import HttpResponse
# from django.utils import timezone
# import openpyxl
# from .models import CAMSTransaction, CAMSSummary


# def download_previous_month(request):
#     # Compute previous month window as dates
#     today = timezone.now().date()
#     first_day_current = today.replace(day=1)
#     last_day_previous = first_day_current - timedelta(days=1)

#     start_date = last_day_previous.replace(day=1)   # e.g., 2023-07-01
#     end_date   = last_day_previous                  # e.g., 2023-07-31
#     filename = f"CAMS_Previous_Month_{last_day_previous.strftime('%Y_%m')}.xlsx"

#     # Generate and return the Excel
#     return generate_excel(start_date, end_date, filename)


# def _to_number(x):
#     if x is None:
#         return None
#     try:
#         return float(x)
#     except Exception:
#         return None


# def _to_excel_date(x):
#     """Return a date/datetime as-is (openpyxl handles it), otherwise string/None."""
#     if x is None:
#         return None
#     if isinstance(x, (date, datetime)):
#         return x
#     return str(x)


# def generate_excel(start_date, end_date, filename):
#     """
#     Builds an Excel with:
#       - Sheet 1: Transactions (filtered by previous month)
#       - Sheet 2: Portfolio Summary (filtered by previous month using nav_date)
#     Uses __year/__month lookups so results are not missed due to time components.
#     """

#     # Month/year for the previous month (derived from start_date which is 1st of prev month)
#     prev_month = start_date.month
#     prev_year  = start_date.year

#     # -------- Transactions (by transaction_date) --------
#     transactions = CAMSTransaction.objects.filter(
#         transaction_date__year=prev_year,
#         transaction_date__month=prev_month
#     ).order_by('entity_name', 'folio_number', 'transaction_date')

#     # -------- Portfolio Summaries (by nav_date) --------
#     summaries = CAMSSummary.objects.filter(
#         nav_date__year=prev_year,
#         nav_date__month=prev_month
#     ).order_by('entity_name', 'folio_number', 'nav_date')

#     # Debug logs in server console
#     print(
#         f"[Excel] Period {start_date} to {end_date} | "
#         f"Tx count={transactions.count()} | Summary count={summaries.count()}"
#     )

#     # -------- Create Workbook --------
#     wb = openpyxl.Workbook()
#     ws_tx = wb.active
#     ws_tx.title = "Transactions"

#     # Headers
#     trans_headers = [
#         'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
#         'Transaction Date', 'Transaction Description', 'Value', 'Stamp Duty',
#         'Units', 'NAV', 'Unit Balance'
#     ]
#     for c, h in enumerate(trans_headers, 1):
#         ws_tx.cell(row=1, column=c, value=h)

#     # Rows
#     for r, t in enumerate(transactions, 2):
#         ws_tx.cell(row=r, column=1,  value=t.entity_name)
#         ws_tx.cell(row=r, column=2,  value=t.amc_name)
#         ws_tx.cell(row=r, column=3,  value=t.folio_number)
#         ws_tx.cell(row=r, column=4,  value=t.isin)
#         ws_tx.cell(row=r, column=5,  value=t.scheme_name)
#         ws_tx.cell(row=r, column=6,  value=_to_excel_date(t.transaction_date))
#         ws_tx.cell(row=r, column=7,  value=t.transaction_description)
#         ws_tx.cell(row=r, column=8,  value=_to_number(t.value))
#         ws_tx.cell(row=r, column=9,  value=_to_number(t.stamp_duty))
#         ws_tx.cell(row=r, column=10, value=_to_number(t.units))
#         ws_tx.cell(row=r, column=11, value=_to_number(t.nav))
#         ws_tx.cell(row=r, column=12, value=_to_number(t.unit_balance))

#     # If no rows, leave a friendly note so the sheet isn’t just “blank”
#     if transactions.count() == 0:
#         ws_tx.cell(row=2, column=1, value=f"No transactions found for {start_date:%b %Y}")

#     # -------- Portfolio Summary Sheet --------
#     ws_sum = wb.create_sheet(title="Portfolio Summary")

#     summary_headers = [
#         'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
#         'Closing Unit Balance', 'NAV', 'Total Cost Value', 'Market Value', 'NAV Date'
#     ]
#     for c, h in enumerate(summary_headers, 1):
#         ws_sum.cell(row=1, column=c, value=h)

#     for r, s in enumerate(summaries, 2):
#         ws_sum.cell(row=r, column=1,  value=s.entity_name)
#         ws_sum.cell(row=r, column=2,  value=s.amc_name)
#         ws_sum.cell(row=r, column=3,  value=s.folio_number)
#         ws_sum.cell(row=r, column=4,  value=s.isin)
#         ws_sum.cell(row=r, column=5,  value=s.scheme_name)
#         ws_sum.cell(row=r, column=6,  value=_to_number(s.closing_unit_balance))
#         ws_sum.cell(row=r, column=7,  value=_to_number(s.nav))
#         ws_sum.cell(row=r, column=8,  value=_to_number(s.total_cost_value))
#         ws_sum.cell(row=r, column=9,  value=_to_number(s.market_value))
#         ws_sum.cell(row=r, column=10, value=_to_excel_date(s.nav_date))

#     if summaries.count() == 0:
#         ws_sum.cell(row=2, column=1, value=f"No summaries found for {start_date:%b %Y}")

#     # Optional niceties: freeze header rows
#     ws_tx.freeze_panes = "A2"
#     ws_sum.freeze_panes = "A2"

#     # Save to memory
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     # Response
#     resp = HttpResponse(
#         output.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     resp['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return resp








# # ====================================   working code ===========================================
# import io
# from datetime import timedelta, date, datetime
# from django.http import HttpResponse
# from django.utils import timezone
# import openpyxl
# from .models import CAMSTransaction, CAMSSummary


# def get_month_year_for_current_and_previous():
#     today = timezone.now().date()

#     if today.day == 1:
#         # On 1st: current month = previous calendar month
#         first_day_current = today.replace(day=1)
#         last_day_previous = first_day_current - timedelta(days=1)

#         current_month = last_day_previous.month
#         current_year = last_day_previous.year

#         # Previous month = one before current
#         prev_last_day = last_day_previous.replace(day=1) - timedelta(days=1)
#         previous_month = prev_last_day.month
#         previous_year = prev_last_day.year
#     else:
#         # From 2nd onwards: normal behavior
#         current_month = today.month
#         current_year = today.year

#         # Previous month = just before current
#         first_day_current = today.replace(day=1)
#         last_day_previous = first_day_current - timedelta(days=1)
#         previous_month = last_day_previous.month
#         previous_year = last_day_previous.year

#     return (current_year, current_month), (previous_year, previous_month)


# def month_date_range(year, month):
#     """Return (start_date, end_date) for a given year/month."""
#     start_date = date(year, month, 1)
#     # trick: go to 28th, add 4 days → definitely next month
#     next_month = start_date.replace(day=28) + timedelta(days=4)
#     end_date = next_month - timedelta(days=next_month.day)
#     return start_date, end_date


# def download_current_month(request):
#     (cur_year, cur_month), _ = get_month_year_for_current_and_previous()
#     start_date, end_date = month_date_range(cur_year, cur_month)

#     filename = f"CAMS_Current_Month_{cur_year}_{cur_month:02d}.xlsx"
#     return generate_excel(start_date, end_date, filename)


# def download_previous_month(request):
#     _, (prev_year, prev_month) = get_month_year_for_current_and_previous()
#     start_date, end_date = month_date_range(prev_year, prev_month)

#     filename = f"CAMS_Previous_Month_{prev_year}_{prev_month:02d}.xlsx"
#     return generate_excel(start_date, end_date, filename)


# def _to_number(x):
#     if x is None:
#         return None
#     try:
#         return float(x)
#     except Exception:
#         return None


# def _to_excel_date(x):
#     """Return a date/datetime as-is (openpyxl handles it), otherwise string/None."""
#     if x is None:
#         return None
#     if isinstance(x, (date, datetime)):
#         return x
#     return str(x)


# def generate_excel(start_date, end_date, filename):
#     """
#     Builds an Excel with:
#       - Sheet 1: Transactions (filtered by transaction_date)
#       - Sheet 2: Portfolio Summary (filtered by nav_date)
#     """

#     # -------- Transactions (by transaction_date) --------
#     transactions = CAMSTransaction.objects.filter(
#         transaction_date__range=[start_date, end_date]
#     ).order_by('entity_name', 'folio_number', 'transaction_date')

#     # -------- Portfolio Summaries (by nav_date) --------
#     summaries = CAMSSummary.objects.filter(
#         nav_date__range=[start_date, end_date]
#     ).order_by('entity_name', 'folio_number', 'nav_date')

#     # Debug log
#     print(
#         f"[Excel] Period {start_date} to {end_date} | "
#         f"Tx count={transactions.count()} | Summary count={summaries.count()}"
#     )

#     # -------- Create Workbook --------
#     wb = openpyxl.Workbook()
#     ws_tx = wb.active
#     ws_tx.title = "Transactions"

#     # Headers
#     trans_headers = [
#         'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
#         'Transaction Date', 'Transaction Description', 'Value', 'Stamp Duty',
#         'Units', 'NAV', 'Unit Balance'
#     ]
#     for c, h in enumerate(trans_headers, 1):
#         ws_tx.cell(row=1, column=c, value=h)

#     # Rows
#     for r, t in enumerate(transactions, 2):
#         ws_tx.cell(row=r, column=1, value=t.entity_name)
#         ws_tx.cell(row=r, column=2, value=t.amc_name)
#         ws_tx.cell(row=r, column=3, value=t.folio_number)
#         ws_tx.cell(row=r, column=4, value=t.isin)
#         ws_tx.cell(row=r, column=5, value=t.scheme_name)
#         ws_tx.cell(row=r, column=6, value=_to_excel_date(t.transaction_date))
#         ws_tx.cell(row=r, column=7, value=t.transaction_description)
#         ws_tx.cell(row=r, column=8, value=_to_number(t.value))
#         ws_tx.cell(row=r, column=9, value=_to_number(t.stamp_duty))
#         ws_tx.cell(row=r, column=10, value=_to_number(t.units))
#         ws_tx.cell(row=r, column=11, value=_to_number(t.nav))
#         ws_tx.cell(row=r, column=12, value=_to_number(t.unit_balance))

#     if transactions.count() == 0:
#         ws_tx.cell(row=2, column=1, value=f"No transactions found for {start_date:%b %Y}")

#     # -------- Portfolio Summary Sheet --------
#     ws_sum = wb.create_sheet(title="Portfolio Summary")

#     summary_headers = [
#         'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
#         'Closing Unit Balance', 'NAV', 'Total Cost Value', 'Market Value', 'NAV Date'
#     ]
#     for c, h in enumerate(summary_headers, 1):
#         ws_sum.cell(row=1, column=c, value=h)

#     for r, s in enumerate(summaries, 2):
#         ws_sum.cell(row=r, column=1, value=s.entity_name)
#         ws_sum.cell(row=r, column=2, value=s.amc_name)
#         ws_sum.cell(row=r, column=3, value=s.folio_number)
#         ws_sum.cell(row=r, column=4, value=s.isin)
#         ws_sum.cell(row=r, column=5, value=s.scheme_name)
#         ws_sum.cell(row=r, column=6, value=_to_number(s.closing_unit_balance))
#         ws_sum.cell(row=r, column=7, value=_to_number(s.nav))
#         ws_sum.cell(row=r, column=8, value=_to_number(s.total_cost_value))
#         ws_sum.cell(row=r, column=9, value=_to_number(s.market_value))
#         ws_sum.cell(row=r, column=10, value=_to_excel_date(s.nav_date))

#     if summaries.count() == 0:
#         ws_sum.cell(row=2, column=1, value=f"No summaries found for {start_date:%b %Y}")

#     # Freeze header rows
#     ws_tx.freeze_panes = "A2"
#     ws_sum.freeze_panes = "A2"

#     # Save to memory
#     output = io.BytesIO()
#     wb.save(output)
#     output.seek(0)

#     # Response
#     resp = HttpResponse(
#         output.getvalue(),
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     resp['Content-Disposition'] = f'attachment; filename="{filename}"'
#     return resp


# # ====================================   working code ===========================================


import io
from datetime import timedelta, date, datetime
from django.http import HttpResponse
import openpyxl
from .models import CAMSTransaction, CAMSSummary


# --- HARDCODED "today" ---
HARDCODED_TODAY = date(2025, 9, 1)


def get_month_year_for_current_and_previous():
    today = HARDCODED_TODAY  # fixed as 2nd Sep 2025

    if today.day == 1:
        # On 1st: current month = previous calendar month
        first_day_current = today.replace(day=1)
        last_day_previous = first_day_current - timedelta(days=1)

        current_month = last_day_previous.month
        current_year = last_day_previous.year

        # Previous month = one before current
        prev_last_day = last_day_previous.replace(day=1) - timedelta(days=1)
        previous_month = prev_last_day.month
        previous_year = prev_last_day.year
    else:
        # From 2nd onwards: normal behavior
        current_month = today.month
        current_year = today.year

        # Previous month = just before current
        first_day_current = today.replace(day=1)
        last_day_previous = first_day_current - timedelta(days=1)
        previous_month = last_day_previous.month
        previous_year = last_day_previous.year

    return (current_year, current_month), (previous_year, previous_month)


def month_date_range(year, month):
    """Return (start_date, end_date) for a given year/month."""
    start_date = date(year, month, 1)
    # trick: go to 28th, add 4 days → definitely next month
    next_month = start_date.replace(day=28) + timedelta(days=4)
    end_date = next_month - timedelta(days=next_month.day)
    return start_date, end_date


def download_current_month(request):
    (cur_year, cur_month), _ = get_month_year_for_current_and_previous()
    start_date, end_date = month_date_range(cur_year, cur_month)

    filename = f"CAMS_Current_Month_{cur_year}_{cur_month:02d}.xlsx"
    return generate_excel(start_date, end_date, filename)


def download_previous_month(request):
    _, (prev_year, prev_month) = get_month_year_for_current_and_previous()
    start_date, end_date = month_date_range(prev_year, prev_month)

    filename = f"CAMS_Previous_Month_{prev_year}_{prev_month:02d}.xlsx"
    return generate_excel(start_date, end_date, filename)


def _to_number(x):
    if x is None:
        return None
    try:
        return float(x)
    except Exception:
        return None


def _to_excel_date(x):
    """Return a date/datetime as-is (openpyxl handles it), otherwise string/None."""
    if x is None:
        return None
    if isinstance(x, (date, datetime)):
        return x
    return str(x)


def generate_excel(start_date, end_date, filename):
    """
    Builds an Excel with:
      - Sheet 1: Transactions (filtered by transaction_date)
      - Sheet 2: Portfolio Summary (filtered by nav_date)
    """

    # -------- Transactions (by transaction_date) --------
    transactions = CAMSTransaction.objects.filter(
        transaction_date__range=[start_date, end_date]
    ).order_by('entity_name', 'folio_number', 'transaction_date')

    # -------- Portfolio Summaries (by nav_date) --------
    summaries = CAMSSummary.objects.filter(
        nav_date__range=[start_date, end_date]
    ).order_by('entity_name', 'folio_number', 'nav_date')

    # Debug log
    print(
        f"[Excel] Period {start_date} to {end_date} | "
        f"Tx count={transactions.count()} | Summary count={summaries.count()}"
    )

    # -------- Create Workbook --------
    wb = openpyxl.Workbook()
    ws_tx = wb.active
    ws_tx.title = "Transactions"

    # Headers
    trans_headers = [
        'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
        'Transaction Date', 'Transaction Description', 'Value', 'Stamp Duty',
        'Units', 'NAV', 'Unit Balance'
    ]
    for c, h in enumerate(trans_headers, 1):
        ws_tx.cell(row=1, column=c, value=h)

    # Rows
    for r, t in enumerate(transactions, 2):
        ws_tx.cell(row=r, column=1, value=t.entity_name)
        ws_tx.cell(row=r, column=2, value=t.amc_name)
        ws_tx.cell(row=r, column=3, value=t.folio_number)
        ws_tx.cell(row=r, column=4, value=t.isin)
        ws_tx.cell(row=r, column=5, value=t.scheme_name)
        ws_tx.cell(row=r, column=6, value=_to_excel_date(t.transaction_date))
        ws_tx.cell(row=r, column=7, value=t.transaction_description)
        ws_tx.cell(row=r, column=8, value=_to_number(t.value))
        ws_tx.cell(row=r, column=9, value=_to_number(t.stamp_duty))
        ws_tx.cell(row=r, column=10, value=_to_number(t.units))
        ws_tx.cell(row=r, column=11, value=_to_number(t.nav))
        ws_tx.cell(row=r, column=12, value=_to_number(t.unit_balance))

    if transactions.count() == 0:
        ws_tx.cell(row=2, column=1, value=f"No transactions found for {start_date:%b %Y}")

    # -------- Portfolio Summary Sheet --------
    ws_sum = wb.create_sheet(title="Portfolio Summary")

    summary_headers = [
        'Entity Name', 'AMC Name', 'Folio Number', 'ISIN', 'Scheme Name',
        'Closing Unit Balance', 'NAV', 'Total Cost Value', 'Market Value', 'NAV Date'
    ]
    for c, h in enumerate(summary_headers, 1):
        ws_sum.cell(row=1, column=c, value=h)

    for r, s in enumerate(summaries, 2):
        ws_sum.cell(row=r, column=1, value=s.entity_name)
        ws_sum.cell(row=r, column=2, value=s.amc_name)
        ws_sum.cell(row=r, column=3, value=s.folio_number)
        ws_sum.cell(row=r, column=4, value=s.isin)
        ws_sum.cell(row=r, column=5, value=s.scheme_name)
        ws_sum.cell(row=r, column=6, value=_to_number(s.closing_unit_balance))
        ws_sum.cell(row=r, column=7, value=_to_number(s.nav))
        ws_sum.cell(row=r, column=8, value=_to_number(s.total_cost_value))
        ws_sum.cell(row=r, column=9, value=_to_number(s.market_value))
        ws_sum.cell(row=r, column=10, value=_to_excel_date(s.nav_date))

    if summaries.count() == 0:
        ws_sum.cell(row=2, column=1, value=f"No summaries found for {start_date:%b %Y}")

    # Freeze header rows
    ws_tx.freeze_panes = "A2"
    ws_sum.freeze_panes = "A2"

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Response
    resp = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
